#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
import openpyxl
from datetime import datetime, timedelta, timezone
import sys
from getpass import getpass
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import json
import os


def setup_proxy(username, password, http_proxy, https_proxy):
    """设置代理配置"""
    if username and password:
        # 在代理URL中添加认证信息
        http_proxy_with_auth = http_proxy.replace('http://', f'http://{username}:{password}@')
        https_proxy_with_auth = https_proxy.replace('http://', f'http://{username}:{password}@')
        proxies = {
            'http': http_proxy_with_auth,
            'https': https_proxy_with_auth
        }
    else:
        proxies = {
            'http': http_proxy,
            'https': https_proxy
        }
    return proxies


def read_pubdev_packages(excel_file):
    """从Excel文件读取pub.dev库名称列表"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        packages = []
        
        # 读取第一列的所有非空值（跳过表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 如果第一列有值
                packages.append(str(row[0]).strip())
        
        wb.close()
        return packages
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        sys.exit(1)


def get_package_versions(package_name, proxies):
    """获取pub.dev包的版本信息"""
    url = f"https://pub.dev/api/packages/{package_name}"
    headers = {
        'Accept': 'application/vnd.pub.v2+json'
    }
    
    try:
        response = requests.get(url, headers=headers, proxies=proxies, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"获取包 {package_name} 信息失败: {e}")
        return None


def get_latest_version(package_data, package_name, proxies):
    """获取最新版本信息"""
    if not package_data or 'versions' not in package_data:
        return []
    
    versions = package_data.get('versions', [])
    
    if not versions:
        return []
    
    # 获取最新版本（版本列表的最后一个）
    version_data = versions[-1]
    
    try:
        # 从版本数据中获取版本号和发布时间
        version = version_data.get('version', '')
        publish_time = version_data.get('published', '')
        pubspec = version_data.get('pubspec', {})
        
        # 获取描述和依赖信息
        description = pubspec.get('description', '')
        
        # 获取作者信息
        author_info = pubspec.get('author', '') or pubspec.get('authors', [])
        if isinstance(author_info, list):
            author = author_info[0] if author_info else ''
        else:
            author = str(author_info)
        
        # 获取依赖数量
        dependencies = len(pubspec.get('dependencies', {}))
        
        version_info = {
            'version': version,
            'publish_time': publish_time,
            'description': description,
            'author': author,
            'dependencies': dependencies
        }
        return [version_info]
    except Exception as e:
        # 静默处理错误，不打印
        return []


def scan_single_package(package_name, proxies, lock, progress):
    """扫描单个pub.dev包的版本信息（用于多线程）"""
    try:
        package_data = get_package_versions(package_name, proxies)
        
        if package_data:
            versions = get_latest_version(package_data, package_name, proxies)
            result = versions
            if versions:
                status_msg = f"✓ 找到最新版本: {versions[0]['version']}"
            else:
                status_msg = "✓ 未找到版本"
        else:
            result = None
            status_msg = "✗ 获取失败"
        
        # 线程安全地更新进度
        with lock:
            progress['completed'] += 1
            print(f"[{progress['completed']}/{progress['total']}] {package_name}: {status_msg}")
        
        return package_name, result
    except Exception as e:
        with lock:
            progress['completed'] += 1
            print(f"[{progress['completed']}/{progress['total']}] {package_name}: ✗ 异常: {e}")
        return package_name, None


def write_results_to_excel(results, output_file):
    """将扫描结果写入Excel文件"""
    wb = openpyxl.Workbook()
    
    # 第一个sheet：详细版本信息
    ws1 = wb.active
    ws1.title = "详细版本信息"
    
    # 写入表头
    headers = ['包名', '版本', '发布时间', '描述', '作者', '依赖数量']
    ws1.append(headers)
    
    # 写入数据
    for package_name, versions in results.items():
        if versions is None:
            ws1.append([package_name, '查找失败', '', '', '', ''])
        elif not versions:
            ws1.append([package_name, '未找到版本', '', '', '', ''])
        else:
            for version_info in versions:
                ws1.append([
                    package_name,
                    version_info['version'],
                    version_info['publish_time'],
                    version_info['description'],
                    version_info['author'],
                    version_info['dependencies']
                ])
    
    # 调整列宽
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # 第二个sheet：统计信息
    ws2 = wb.create_sheet(title="版本统计")
    ws2.append(['库名', '最新版本'])
    
    # 写入统计数据
    for package_name, versions in results.items():
        if versions is None:
            ws2.append([package_name, '查找失败'])
        elif not versions:
            ws2.append([package_name, '未找到版本'])
        else:
            ws2.append([package_name, versions[0]['version']])
    
    # 调整统计页列宽
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    print(f"\n扫描结果已保存到: {output_file}")


def main():
    print("=" * 60)
    print("Pub.dev库版本扫描工具")
    print("=" * 60)
    
    # 1. 获取代理配置
    print("\n请输入代理配置（如果不需要代理，直接按回车跳过）:")
    proxy_username = input("代理用户名: ").strip()
    proxy_password = ""
    if proxy_username:
        proxy_password = getpass("代理密码: ")
    
    http_proxy = input("HTTP Proxy (例如: http://proxy.example.com:8080): ").strip()
    https_proxy = input("HTTPS Proxy (例如: http://proxy.example.com:8080): ").strip()
    
    proxies = None
    if http_proxy or https_proxy:
        proxies = setup_proxy(proxy_username, proxy_password, http_proxy, https_proxy)
        print("✓ 代理配置完成")
    
    # 2. 获取输入文件路径
    print("\n请输入包含pub.dev库名称的Excel文件路径:")
    input_file = input("Excel文件路径: ").strip()
    if not input_file:
        print("错误: 必须提供Excel文件路径")
        sys.exit(1)
    
    # 3. 读取pub.dev包列表
    print(f"\n正在读取Excel文件: {input_file}")
    packages = read_pubdev_packages(input_file)
    print(f"✓ 共读取到 {len(packages)} 个pub.dev包")
    
    # 4. 扫描每个包的版本信息
    print("\n开始扫描pub.dev包版本信息...")
    print("使用多线程并发扫描，请稍候...\n")
    
    results = {}
    lock = threading.Lock()
    progress = {'completed': 0, 'total': len(packages)}
    
    # 使用线程池并发扫描，max_workers控制并发数
    with ThreadPoolExecutor(max_workers=15) as executor:
        # 提交所有任务
        future_to_package = {
            executor.submit(scan_single_package, package, proxies, lock, progress): package 
            for package in packages
        }
        
        # 收集结果
        for future in as_completed(future_to_package):
            try:
                package_name, result = future.result()
                results[package_name] = result
            except Exception as e:
                package_name = future_to_package[future]
                results[package_name] = None
                print(f"处理 {package_name} 时发生异常: {e}")
    
    # 6. 输出结果到Excel
    print("\n正在生成结果文件...")
    output_file = input_file.replace('.xlsx', '-扫描结果.xlsx')
    if output_file == input_file:
        output_file = input_file.replace('.xlsx', '') + '-扫描结果.xlsx'
    
    write_results_to_excel(results, output_file)
    
    # 统计信息
    success_count = sum(1 for versions in results.values() if versions is not None and versions)
    failed_count = sum(1 for versions in results.values() if versions is None)
    not_found_count = sum(1 for versions in results.values() if versions is not None and not versions)
    print("\n" + "=" * 60)
    print("扫描完成!")
    print(f"共扫描 {len(packages)} 个pub.dev包")
    print(f"成功获取 {success_count} 个包的最新版本")
    print(f"查找失败 {failed_count} 个pub.dev包")
    print(f"未找到版本 {not_found_count} 个pub.dev包")
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n程序执行出错: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
